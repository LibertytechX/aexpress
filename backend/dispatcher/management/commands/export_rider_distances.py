from __future__ import annotations

import datetime
from decimal import Decimal
from typing import Any, List, Optional

from django.core.management.base import BaseCommand, CommandParser
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from dispatcher.models import Rider, VehicleAsset, VehicleTracking, VehicleReassignment


def get_assigned_asset_on_date(rider: Rider, date: datetime.date) -> Optional[VehicleAsset]:
    """Retrieve the vehicle asset assigned to the rider on a specific date.

    Args:
        rider: The Rider instance to check.
        date: The date to look up the assigned asset for.

    Returns:
        The VehicleAsset assigned on that date, or None if none.
    """
    tz = timezone.get_current_timezone()
    end_of_day = timezone.make_aware(
        datetime.datetime.combine(date, datetime.time.max), tz
    )

    # Check if there is any reassignment to the rider before the end of the day
    last_assign_to = (
        VehicleReassignment.objects.filter(to_rider=rider, created_at__lte=end_of_day)
        .order_by("-created_at")
        .first()
    )

    # Check if there is any reassignment from the rider after that assignment but before the end of the day
    if last_assign_to:
        last_assign_from = VehicleReassignment.objects.filter(
            from_rider=rider,
            vehicle_asset=last_assign_to.vehicle_asset,
            created_at__gt=last_assign_to.created_at,
            created_at__lte=end_of_day,
        ).exists()
        if not last_assign_from:
            return last_assign_to.vehicle_asset

    # Fallback to current vehicle asset
    return rider.vehicle_asset


def get_distance_for_asset_on_date(asset_id: Any, date: datetime.date) -> float:
    """Calculate the distance travelled by an asset on a specific date from telemetry tracking.

    Args:
        asset_id: The ID of the VehicleAsset.
        date: The date to query.

    Returns:
        The distance in kilometers (float).
    """
    tz = timezone.get_current_timezone()
    start = timezone.make_aware(
        datetime.datetime.combine(date, datetime.time.min), tz
    )
    end = timezone.make_aware(
        datetime.datetime.combine(date + datetime.timedelta(days=1), datetime.time.min), tz
    )

    trackings = VehicleTracking.objects.filter(
        vehicle_asset_id=asset_id,
        created_at__gte=start,
        created_at__lt=end,
        travelled__isnull=False,
    ).order_by("created_at")

    if trackings.exists():
        first_entry = trackings.first()
        last_entry = trackings.last()
        if (
            first_entry is not None
            and last_entry is not None
            and first_entry.travelled is not None
            and last_entry.travelled is not None
        ):
            distance = float(last_entry.travelled) - float(first_entry.travelled)
            return max(0.00, round(distance, 2))
    return 0.00


class Command(BaseCommand):
    """Management command to export rider distance covered data to Excel."""

    help: str = (
        "Generate rider distance covered report in an Excel sheet for a specified date range."
    )

    def add_arguments(self, parser: CommandParser) -> None:
        """Add arguments to the command parser.

        Args:
            parser: The command line argument parser.
        """
        parser.add_argument(
            "--start-date",
            dest="start_date",
            default=None,
            help="Start date in YYYY-MM-DD format (defaults to the 1st day of the current month).",
        )
        parser.add_argument(
            "--end-date",
            dest="end_date",
            default=None,
            help="End date in YYYY-MM-DD format (defaults to today).",
        )
        parser.add_argument(
            "--output",
            dest="output",
            default="rider_distances.xlsx",
            help="Filename/path for the output Excel file (defaults to 'rider_distances.xlsx').",
        )
        parser.add_argument(
            "--all-riders",
            dest="all_riders",
            action="store_true",
            help="Include riders who do not have an assigned vehicle asset.",
        )

    def handle(self, *args: Any, **options: Any) -> None:
        """Execute the command logic.

        Args:
            *args: Variable length argument list.
            **options: Arbitrary keyword arguments.
        """
        start_date_str: Optional[str] = options.get("start_date")
        end_date_str: Optional[str] = options.get("end_date")
        output_file: str = options.get("output", "rider_distances.xlsx")
        all_riders: bool = bool(options.get("all_riders"))

        today = timezone.localdate()

        # Resolve start and end dates
        if start_date_str:
            start_date = datetime.date.fromisoformat(start_date_str)
        else:
            start_date = today.replace(day=1)

        if end_date_str:
            end_date = datetime.date.fromisoformat(end_date_str)
        else:
            end_date = today

        if start_date > end_date:
            self.stderr.write(
                self.style.ERROR(
                    f"Start date ({start_date}) cannot be after end date ({end_date})."
                )
            )
            return

        self.stdout.write(
            f"Generating report from {start_date} to {end_date}..."
        )

        # Generate list of dates in descending order
        delta_days = (end_date - start_date).days
        dates: List[datetime.date] = [
            end_date - datetime.timedelta(days=i) for i in range(delta_days + 1)
        ]

        # Get relevant riders
        if all_riders:
            riders = Rider.objects.select_related("user", "vehicle_asset").all()
        else:
            riders = Rider.objects.select_related("user", "vehicle_asset").filter(
                vehicle_asset__isnull=False
            )

        riders = riders.order_by("user__contact_name")

        # Initialize openpyxl workbook
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Rider Distances"

        # Ensure gridlines are visible
        ws.views.sheetView[0].showGridLines = True

        # Columns configuration
        headers = [
            "Driver Name",
            "Vehicle Plate",
            "",  # Empty column matching screenshot
            "Date",
            "Distance Covered (km)",
        ]
        ws.append(headers)

        # Define premium styled theme (Liberty Corporate Theme)
        navy_fill = PatternFill(
            start_color="1B365D", end_color="1B365D", fill_type="solid"
        )
        zebra_fill = PatternFill(
            start_color="F4F6F9", end_color="F4F6F9", fill_type="solid"
        )
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=11, color="000000")
        
        thin_border = Border(
            left=Side(style="thin", color="E0E0E0"),
            right=Side(style="thin", color="E0E0E0"),
            top=Side(style="thin", color="E0E0E0"),
            bottom=Side(style="thin", color="E0E0E0"),
        )

        # Apply header styling
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        ws.row_dimensions[1].height = 28

        # Populate rows
        row_idx = 2
        for rider in riders:
            rider_name = rider.user.contact_name or rider.user.phone or "Unknown Rider"
            
            # Keep zebra fill state per driver to make it visually clean
            for day in dates:
                asset = get_assigned_asset_on_date(rider, day)
                plate = asset.plate_number if asset else ""
                
                # Fetch distance
                distance = 0.00
                if asset:
                    distance = get_distance_for_asset_on_date(asset.id, day)

                # Format date to match M/D/YYYY
                date_str = f"{day.month}/{day.day}/{day.year}"

                # Append data
                row_data = [
                    rider_name,
                    plate,
                    "",
                    date_str,
                    distance,
                ]
                ws.append(row_data)

                # Style data cells
                for col_num in range(1, 6):
                    cell = ws.cell(row=row_idx, column=col_num)
                    cell.font = data_font
                    cell.border = thin_border
                    
                    # Zebra striping (alternate background rows)
                    if row_idx % 2 == 0:
                        cell.fill = zebra_fill

                    # Set alignments
                    if col_num in (1, 2):
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif col_num in (3, 4):
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_num == 5:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = "#,##0.00"

                ws.row_dimensions[row_idx].height = 20
                row_idx += 1

        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            # Give a bit of padding, unless it's the blank column C
            if col_letter == "C":
                ws.column_dimensions[col_letter].width = 5
            else:
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(output_file)
        self.stdout.write(
            self.style.SUCCESS(
                f"Successfully exported rider distance data to {output_file}"
            )
        )
